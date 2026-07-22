"""Glossary management routes for PrionVault.

Handles glossary CRUD, summary improvement, and statistics.
Routes registered as side-effect import at bottom of routes.py.
"""
import logging
import threading
from io import BytesIO
from datetime import datetime
from flask import jsonify, request, Response, current_app, send_file
from sqlalchemy import text as sql_text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.decorators import admin_required, login_required
from database.config import db
from . import prionvault_bp

logger = logging.getLogger(__name__)

# Track active batch processing
_batch_state = {"status": None, "error": None, "queued": 0, "processed": 0}


# ── Glossary stats & dashboard ─────────────────────────────────────────────
@prionvault_bp.route("/api/glossary/stats", methods=["GET"])
@admin_required
def api_glossary_stats():
    """Get comprehensive glossary statistics for dashboard."""
    from .services import summary_improver

    try:
        stats = summary_improver.get_improvement_stats()
        return jsonify(stats)
    except Exception as e:
        logger.exception("Failed to fetch glossary stats")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/stats/detailed", methods=["GET"])
@admin_required
def api_glossary_stats_detailed():
    """Get detailed glossary review status breakdown."""
    from .services import glossary_manager

    current_version = glossary_manager.get_current_glossary_version()

    try:
        # Check if summary_improvement_log table exists using information_schema
        table_exists = False
        try:
            with db.engine.connect() as check_conn:
                result = check_conn.execute(sql_text("""
                    SELECT EXISTS (
                        SELECT 1 FROM information_schema.tables
                        WHERE table_name = 'summary_improvement_log'
                    )
                """)).scalar()
                table_exists = bool(result)
        except Exception as e:
            logger.warning(f"Failed to check table existence: {e}")
            table_exists = False

        with db.engine.connect() as conn:
            # Pending: unreviewed summaries (ai_summary_glossary_version IS NULL)
            pending = conn.execute(sql_text("""
                SELECT COUNT(*) FROM articles
                WHERE summary_ai IS NOT NULL
                  AND ai_summary_glossary_version IS NULL
                  AND char_length(summary_ai) > 50
            """)).scalar() or 0

            # Reviewed stats: only query if table exists
            if table_exists:
                # Reviewed with changes: articles with improvement log entries
                reviewed_with_changes = conn.execute(sql_text("""
                    SELECT COUNT(DISTINCT article_id)
                    FROM summary_improvement_log
                    WHERE changes_count > 0 AND dry_run = FALSE
                """)).scalar() or 0

                # Reviewed without changes: articles with improvement log but no changes
                reviewed_without_changes = conn.execute(sql_text("""
                    SELECT COUNT(DISTINCT article_id)
                    FROM summary_improvement_log
                    WHERE changes_count = 0 AND dry_run = FALSE
                """)).scalar() or 0

                total_reviewed = reviewed_with_changes + reviewed_without_changes
            else:
                reviewed_with_changes = 0
                reviewed_without_changes = 0
                total_reviewed = 0

        return jsonify({
            "pending": int(pending),
            "reviewed_with_changes": int(reviewed_with_changes),
            "reviewed_without_changes": int(reviewed_without_changes),
            "total_reviewed": int(total_reviewed),
            "current_glossary_version": current_version,
        })
    except Exception as e:
        logger.exception("Failed to fetch glossary detailed stats")
        return jsonify({"error": str(e)[:300]}), 500


# ── Unreviewed summaries (glossary_version IS NULL) ──────────────────────
@prionvault_bp.route("/api/glossary/unreviewed", methods=["GET"])
@admin_required
def api_glossary_unreviewed():
    """Get articles with unreviewed AI summaries (ai_summary_glossary_version IS NULL)."""
    limit = max(1, min(100, request.args.get("limit", 50, type=int)))
    offset = max(0, request.args.get("offset", 0, type=int))

    try:
        with db.engine.connect() as conn:
            rows = conn.execute(sql_text("""
                SELECT id::text, title, authors, year, summary_ai,
                       char_length(summary_ai) as summary_length,
                       created_at, updated_at
                FROM articles
                WHERE summary_ai IS NOT NULL
                  AND ai_summary_glossary_version IS NULL
                  AND char_length(summary_ai) > 50
                ORDER BY updated_at DESC
                LIMIT :lim OFFSET :off
            """), {"lim": limit, "off": offset}).mappings().all()

            total = conn.execute(sql_text("""
                SELECT COUNT(*) FROM articles
                WHERE summary_ai IS NOT NULL AND ai_summary_glossary_version IS NULL
            """)).scalar() or 0

        return jsonify({
            "articles": [dict(r) for r in rows],
            "total": int(total),
            "limit": limit,
            "offset": offset,
            "has_more": (offset + limit) < total,
        })
    except Exception as e:
        logger.exception(f"Failed to fetch unreviewed summaries: {e}")
        return jsonify({"error": str(e)[:300]}), 500


# ── Outdated summaries (glossary_version < current) ──────────────────────
@prionvault_bp.route("/api/glossary/outdated", methods=["GET"])
@admin_required
def api_glossary_outdated():
    """Get articles improved with older glossary versions."""
    from .services import glossary_manager

    limit = max(1, min(100, request.args.get("limit", 50, type=int)))
    offset = max(0, request.args.get("offset", 0, type=int))

    try:
        current_version = glossary_manager.get_current_glossary_version()

        # Check if summary_improvement_log table exists using information_schema
        table_exists = False
        try:
            with db.engine.connect() as check_conn:
                result = check_conn.execute(sql_text("""
                    SELECT EXISTS (
                        SELECT 1 FROM information_schema.tables
                        WHERE table_name = 'summary_improvement_log'
                    )
                """)).scalar()
                table_exists = bool(result)
        except Exception as e:
            logger.warning(f"Failed to check table existence: {e}")
            table_exists = False

        # If table doesn't exist, return empty results
        if not table_exists:
            return jsonify({
                "articles": [],
                "current_glossary_version": current_version,
                "total_outdated": 0,
                "limit": limit,
                "offset": offset,
                "has_more": False,
            })

        with db.engine.connect() as conn:
            rows = conn.execute(sql_text("""
                SELECT
                  a.id::text,
                  a.title,
                  a.authors,
                  a.year,
                  a.summary_ai,
                  sil.glossary_version_used,
                  sil.improved_at,
                  sil.changes_count,
                  char_length(a.summary_ai) as summary_length
                FROM summary_improvement_log sil
                JOIN articles a ON a.id = sil.article_id
                WHERE sil.glossary_version_used < :current
                  AND sil.dry_run = FALSE
                ORDER BY sil.glossary_version_used ASC, sil.improved_at DESC
                LIMIT :lim OFFSET :off
            """), {
                "current": current_version,
                "lim": limit,
                "off": offset,
            }).mappings().all()

            total = conn.execute(sql_text("""
                SELECT COUNT(*) FROM summary_improvement_log
                WHERE glossary_version_used < :current
                  AND dry_run = FALSE
            """), {"current": current_version}).scalar() or 0

        return jsonify({
            "articles": [dict(r) for r in rows],
            "current_glossary_version": current_version,
            "total_outdated": int(total),
            "limit": limit,
            "offset": offset,
            "has_more": (offset + limit) < total,
        })
    except Exception as e:
        logger.exception(f"Failed to fetch outdated summaries: {e}")
        return jsonify({"error": str(e)[:300]}), 500


# ── Improvement log ────────────────────────────────────────────────────────
@prionvault_bp.route("/api/glossary/log", methods=["GET"])
@admin_required
def api_glossary_log():
    """Get detailed improvement history, optionally filtered by batch."""
    from .services import summary_improver

    batch_id = request.args.get("batch_id")
    limit = max(1, min(100, request.args.get("limit", 50, type=int)))
    offset = max(0, request.args.get("offset", 0, type=int))

    result = summary_improver.get_improvement_log(
        batch_id=batch_id,
        limit=limit,
        offset=offset,
    )
    return jsonify(result)


# ── Batch improve summaries ────────────────────────────────────────────────
@prionvault_bp.route("/api/glossary/improve-next", methods=["POST"])
@admin_required
def api_glossary_improve_next():
    """Improve the next N unreviewed summaries with glossary."""
    from .services import summary_improver, glossary_manager

    data = request.get_json(force=True, silent=True) or {}
    count = data.get("count", 100)  # Default 100, can be 100, 500, or "all"
    dry_run = data.get("dry_run", False)

    # Validate count
    if count == "all":
        limit = 10000  # Get up to 10k (likely all unreviewed)
    elif isinstance(count, int):
        limit = max(1, min(10000, count))
    else:
        return jsonify({"error": "count must be an integer or 'all'"}), 400

    try:
        # Fetch unreviewed articles
        with db.engine.connect() as conn:
            article_ids = conn.execute(sql_text("""
                SELECT id::text FROM articles
                WHERE summary_ai IS NOT NULL
                  AND ai_summary_glossary_version IS NULL
                  AND char_length(summary_ai) > 50
                ORDER BY updated_at DESC
                LIMIT :lim
            """), {"lim": limit}).scalars().all()

        if not article_ids:
            return jsonify({
                "ok": True,
                "queued": 0,
                "message": "No unreviewed summaries found",
                "dry_run": dry_run,
            })

        # Fetch current glossary
        glossary_context = glossary_manager.get_glossary_context()
        glossary_version = glossary_manager.get_current_glossary_version()
        if not glossary_context:
            return jsonify({"error": "No glossary terms available"}), 400

        # Reset batch state
        _batch_state["status"] = "processing"
        _batch_state["error"] = None
        _batch_state["queued"] = len(article_ids)
        _batch_state["processed"] = 0
        _batch_state["batch_id"] = None

        # Run batch in background
        def _run():
            logger.info(f"🚀 THREAD STARTED: Batch thread beginning execution")
            try:
                logger.info(f"📋 Calling batch_improve_summaries with {len(article_ids)} articles")
                def progress_callback(count, article_title="", status="processing"):
                    _batch_state.update({
                        "processed": count,
                        "current_article": article_title,
                        "current_status": status,
                    })
                    logger.info(f"[PROGRESS] {count}/{len(article_ids)}: {article_title} - {status}")

                result = summary_improver.batch_improve_summaries(
                    article_ids=article_ids,
                    glossary_context=glossary_context,
                    glossary_version=glossary_version,
                    dry_run=dry_run,
                    progress_callback=progress_callback,
                )
                logger.info(f"✅ batch_improve_summaries returned successfully")
                _batch_state["status"] = "completed"
                _batch_state["processed"] = result.get("processed", 0)
                _batch_state["batch_id"] = result.get("batch_id")
                logger.info(f"🏁 Batch completed: {result}")
            except Exception as exc:
                logger.exception("❌ Batch improvement failed: %s", exc)
                _batch_state["status"] = "error"
                _batch_state["error"] = str(exc)[:200]
            finally:
                logger.info(f"🔚 THREAD ENDED: Final state = {_batch_state}")

        threading.Thread(target=_run, name="pv-glossary-batch", daemon=True).start()

        return jsonify({
            "ok": True,
            "queued": len(article_ids),
            "dry_run": dry_run,
            "glossary_version": glossary_version,
            "message": f"Queued {len(article_ids)} articles for improvement"
        })
    except Exception as e:
        logger.exception("Failed to queue improvement batch")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/batch-status", methods=["GET"])
@admin_required
def api_glossary_batch_status():
    """Get current batch processing status with real-time progress tracking."""
    return jsonify({
        "status": _batch_state.get("status"),
        "error": _batch_state.get("error"),
        "queued": _batch_state.get("queued", 0),
        "processed": _batch_state.get("processed", 0),
        "batch_id": _batch_state.get("batch_id"),
        "current_article": _batch_state.get("current_article", ""),
        "current_status": _batch_state.get("current_status", ""),
    })


@prionvault_bp.route("/api/glossary/batch-changes/<batch_id>", methods=["GET"])
@admin_required
def api_glossary_batch_changes(batch_id):
    """Get all corrections made in a batch, grouped and counted."""
    try:
        with db.engine.connect() as conn:
            # Get all corrections for this batch, grouped by original→corrected
            rows = conn.execute(sql_text("""
                SELECT
                    scd.original_text,
                    scd.corrected_text,
                    scd.term_en,
                    scd.recommended_es,
                    scd.correction_type,
                    COUNT(*) as change_count,
                    AVG(CAST(scd.confidence_score AS DECIMAL)) as avg_confidence
                FROM summary_correction_detail scd
                JOIN summary_improvement_log sil ON scd.improvement_log_id = sil.id
                WHERE sil.batch_id = :batch_id
                GROUP BY scd.original_text, scd.corrected_text, scd.term_en,
                         scd.recommended_es, scd.correction_type
                ORDER BY change_count DESC, scd.original_text
            """), {"batch_id": batch_id}).mappings().all()

            changes = [dict(r) for r in rows]

            return jsonify({
                "batch_id": batch_id,
                "total_changes": sum(c["change_count"] for c in changes),
                "unique_changes": len(changes),
                "changes": changes,
            })
    except Exception as e:
        logger.exception(f"Failed to fetch batch changes: {e}")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/batch-export/<batch_id>", methods=["GET"])
@admin_required
def api_glossary_batch_export(batch_id):
    """Export batch changes as a formatted Excel file."""
    try:
        with db.engine.connect() as conn:
            # Get batch info
            batch_info = conn.execute(sql_text("""
                SELECT COUNT(DISTINCT article_id) as articles_improved,
                       MAX(improved_at) as completed_at
                FROM summary_improvement_log
                WHERE batch_id = :batch_id
            """), {"batch_id": batch_id}).mappings().first()

            # Get all corrections with article info
            rows = conn.execute(sql_text("""
                SELECT
                    sil.article_id::text,
                    a.title,
                    scd.original_text,
                    scd.corrected_text,
                    scd.term_en,
                    scd.recommended_es,
                    scd.correction_type,
                    scd.confidence_score,
                    sil.improved_at
                FROM summary_correction_detail scd
                JOIN summary_improvement_log sil ON scd.improvement_log_id = sil.id
                JOIN articles a ON sil.article_id = a.id
                WHERE sil.batch_id = :batch_id
                ORDER BY sil.improved_at DESC, a.title, scd.original_text
            """), {"batch_id": batch_id}).mappings().all()

            changes = [dict(r) for r in rows]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cambios"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="366092")
        subheader_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title and metadata
        ws['A1'] = "📊 REPORTE DE CAMBIOS - MEJORA DE RESÚMENES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = left_align

        ws['A2'] = f"Batch ID: {batch_id}"
        ws['A2'].font = Font(size=9, italic=True, color="666666")
        ws['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="666666")

        # Summary stats
        row = 5
        ws[f'A{row}'] = "RESUMEN"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill

        row += 1
        ws[f'A{row}'] = f"Artículos procesados:"
        ws[f'B{row}'] = batch_info['articles_improved'] if batch_info else 0
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = f"Total de cambios:"
        ws[f'B{row}'] = len(changes)
        ws[f'A{row}'].font = Font(bold=True)

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

        # Group changes by article
        from collections import defaultdict
        by_article = defaultdict(list)
        for change in changes:
            key = (change['article_id'], change['title'])
            by_article[key].append(change)

        # Add data grouped by article
        row = 10
        for (article_id, article_title), article_changes in sorted(by_article.items()):
            # Article header with link to PrionVault
            ws[f'A{row}'] = f"📄 {article_title[:60]}"
            ws[f'A{row}'].font = article_font
            ws[f'A{row}'].fill = article_fill

            # Create hyperlink to PrionVault
            prionvault_url = f"/prionvault/?open={article_id}"
            ws[f'A{row}'].hyperlink = prionvault_url
            ws[f'A{row}'].font = Font(bold=True, color="0563C1", underline="single")

            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            # Column headers for this article's changes
            headers = ["Texto Original", "Texto Corregido", "Término EN", "Recomendado ES", "Confianza (%)"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border

            row += 1

            # Data rows for this article
            for change in article_changes:
                confidence = f"{int(change['confidence_score'] * 100)}%" if change['confidence_score'] else "-"

                cells_data = [
                    change['original_text'],
                    change['corrected_text'],
                    change['term_en'] or "-",
                    change['recommended_es'] or "-",
                    confidence,
                ]

                for col_idx, value in enumerate(cells_data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = left_align if col_idx <= 2 else center_align
                    # Alternate row colors for readability
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                row += 1

            # Blank row between articles
            row += 1

        # Freeze panes at header
        ws.freeze_panes = "A11"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as file download
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"batch-{batch_id[:8]}-cambios.xlsx"
        )

    except Exception as e:
        logger.exception(f"Failed to export batch changes: {e}")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/export-all-changes", methods=["GET"])
@admin_required
def api_glossary_export_all_changes():
    """Export all historical changes as a formatted Excel file."""
    try:
        with db.engine.connect() as conn:
            # Get all corrections with batch and article info
            rows = conn.execute(sql_text("""
                SELECT
                    sil.batch_id,
                    sil.article_id,
                    a.title,
                    scd.original_text,
                    scd.corrected_text,
                    scd.term_en,
                    scd.recommended_es,
                    scd.correction_type,
                    scd.confidence_score,
                    sil.improved_at,
                    sil.glossary_version_used
                FROM summary_correction_detail scd
                JOIN summary_improvement_log sil ON scd.improvement_log_id = sil.id
                JOIN articles a ON sil.article_id = a.id
                ORDER BY sil.improved_at DESC, sil.batch_id, a.title
            """)).mappings().all()

            changes = [dict(r) for r in rows]

        # Group by article for better organization
        from collections import defaultdict
        by_article = defaultdict(list)
        for change in changes:
            article_key = (change['article_id'], change['title'])
            by_article[article_key].append(change)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Todos los cambios"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="366092")
        article_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        article_font = Font(bold=True, size=11, color="1F4E78")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title and metadata
        ws['A1'] = "📊 HISTÓRICO COMPLETO DE CAMBIOS - MEJORA DE RESÚMENES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = left_align

        ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=9, italic=True, color="666666")
        ws['A3'] = f"Total de artículos procesados: {len(by_article)}"
        ws['A3'].font = Font(size=9, italic=True, color="666666")
        ws['A4'] = f"Total de cambios registrados: {len(changes)}"
        ws['A4'].font = Font(size=9, italic=True, color="666666")

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15

        current_row = 6

        # Add data grouped by article
        for (article_id, article_title) in sorted(by_article.keys(), key=lambda x: x[1]):
            article_changes = by_article[(article_id, article_title)]

            # Article header with link to PrionVault
            ws[f'A{current_row}'] = f"📄 {article_title[:60]}"
            ws[f'A{current_row}'].font = Font(bold=True, size=11, color="0563C1", underline="single")
            ws[f'A{current_row}'].fill = article_fill

            # Create hyperlink to PrionVault
            prionvault_url = f"/prionvault/?open={article_id}"
            ws[f'A{current_row}'].hyperlink = prionvault_url

            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1

            # Column headers for this article
            headers = ["Texto Original", "Texto Corregido", "Término EN", "Recomendado ES", "Confianza", "Versión", "Fecha"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border

            current_row += 1

            # Data rows for this article
            for change in article_changes:
                confidence = f"{int(change['confidence_score'] * 100)}%" if change['confidence_score'] else "-"
                date_str = change['improved_at'].strftime('%d/%m/%Y') if change['improved_at'] else "-"

                cells_data = [
                    change['original_text'],
                    change['corrected_text'],
                    change['term_en'] or "-",
                    change['recommended_es'] or "-",
                    confidence,
                    f"v{change['glossary_version_used']}",
                    date_str,
                ]

                for col_idx, value in enumerate(cells_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = left_align if col_idx <= 3 else center_align
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                current_row += 1

            # Blank row between articles
            current_row += 1

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as file download
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"cambios-completos-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        )

    except Exception as e:
        logger.exception(f"Failed to export all changes: {e}")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/test-claude", methods=["GET"])
@admin_required
def api_glossary_test_claude():
    """Test if Claude API is working."""
    try:
        from anthropic import Anthropic
        import os

        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY not configured"}), 400

        client = Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=50,
            messages=[{"role": "user", "content": "Say OK"}],
        )

        return jsonify({
            "ok": True,
            "model": response.model,
            "message": response.content[0].text if response.content else "No response",
        })
    except Exception as e:
        logger.exception("Claude test failed")
        return jsonify({"error": str(e)}), 500


@prionvault_bp.route("/api/glossary/test-single", methods=["POST"])
@admin_required
def api_glossary_test_single():
    """Test improving a single article for debugging."""
    from .services import summary_improver, glossary_manager
    import time

    try:
        # Get one unreviewed article
        with db.engine.connect() as conn:
            article = conn.execute(sql_text("""
                SELECT id::text, summary_ai FROM articles
                WHERE summary_ai IS NOT NULL
                  AND ai_summary_glossary_version IS NULL
                  AND char_length(summary_ai) > 50
                LIMIT 1
            """)).first()

        if not article:
            return jsonify({"error": "No unreviewed articles found"}), 400

        article_id, summary = article
        logger.info(f"Testing improvement on {article_id}")

        # Get glossary
        glossary_context = glossary_manager.get_glossary_context()
        if not glossary_context:
            return jsonify({"error": "No glossary available"}), 400

        logger.info("Starting improve_summary call...")
        start = time.time()

        # Test improve (with timeout)
        improvement = summary_improver.improve_summary(
            article_id=article_id,
            original_summary=summary,
            glossary_context=glossary_context,
        )

        elapsed = time.time() - start

        logger.info(f"improve_summary completed in {elapsed:.2f}s")

        return jsonify({
            "ok": True,
            "article_id": article_id,
            "success": improvement.success,
            "original_length": improvement.original_length,
            "improved_length": improvement.improved_length,
            "error": improvement.error,
            "elapsed_seconds": elapsed,
        })

    except Exception as e:
        logger.exception("Test single improvement failed")
        return jsonify({"error": str(e)[:500]}), 500


@prionvault_bp.route("/glossary/quick-test", methods=["GET"])
@admin_required
def glossary_quick_test():
    """Quick test page showing article status and regeneration results."""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>🧪 Prueba Rápida - Glosario</title>
        <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f5f5; padding: 20px; }
        .container { max-width: 1200px; margin: 0 auto; }
        h1 { margin-bottom: 30px; color: #1f2937; }
        .status-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 15px; margin-bottom: 30px; }
        .status-card { background: white; padding: 15px; border-radius: 8px; border-left: 4px solid #3b82f6; }
        .status-card strong { display: block; color: #6b7280; font-size: 12px; margin-bottom: 5px; }
        .status-card .value { font-size: 24px; font-weight: bold; color: #1f2937; }
        .article-list { background: white; border-radius: 8px; overflow: hidden; margin-bottom: 30px; }
        .article-item { border-bottom: 1px solid #e5e7eb; padding: 15px; display: grid; grid-template-columns: 1fr auto auto auto; gap: 15px; align-items: center; }
        .article-item:last-child { border-bottom: none; }
        .article-info { min-width: 0; }
        .article-title { font-weight: 600; color: #1f2937; margin-bottom: 5px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
        .article-meta { font-size: 12px; color: #6b7280; }
        .version-badge { display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: bold; }
        .version-null { background: #fee2e2; color: #991b1b; }
        .version-set { background: #dcfce7; color: #166534; }
        .action-buttons { display: flex; gap: 8px; white-space: nowrap; }
        button { padding: 6px 12px; border: none; border-radius: 4px; cursor: pointer; font-size: 12px; font-weight: 500; }
        .btn-regenerate { background: #f59e0b; color: white; }
        .btn-regenerate:hover { background: #d97706; }
        .btn-regenerate:disabled { background: #d1d5db; cursor: not-allowed; }
        .btn-check { background: #3b82f6; color: white; }
        .btn-check:hover { background: #2563eb; }
        .btn-check:disabled { background: #d1d5db; cursor: not-allowed; }
        .loading { display: inline-block; width: 16px; height: 16px; border: 2px solid #3b82f6; border-top: 2px solid transparent; border-radius: 50%; animation: spin 0.6s linear infinite; }
        @keyframes spin { to { transform: rotate(360deg); } }
        .alert { padding: 15px; border-radius: 8px; margin-bottom: 20px; }
        .alert-info { background: #dbeafe; color: #1e40af; border-left: 4px solid #3b82f6; }
        .alert-success { background: #dcfce7; color: #166534; border-left: 4px solid #16a34a; }
        .alert-error { background: #fee2e2; color: #991b1b; border-left: 4px solid #dc2626; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🧪 Prueba Rápida - Estado de Artículos</h1>

            <div id="alerts"></div>

            <div class="status-grid">
                <div class="status-card">
                    <strong>Artículos Sin Procesar</strong>
                    <div class="value" id="count-unreviewed">—</div>
                </div>
                <div class="status-card">
                    <strong>Versión Glosario</strong>
                    <div class="value" id="glossary-version">—</div>
                </div>
            </div>

            <div class="article-list" id="article-list">
                <div style="padding: 20px; text-align: center; color: #6b7280;">
                    <div class="loading"></div> Cargando artículos...
                </div>
            </div>
        </div>

        <script>
        async function loadArticles() {
            const alerts = document.getElementById('alerts');
            alerts.innerHTML = '';

            try {
                const resp = await fetch('/prionvault/api/glossary/unreviewed?limit=100');
                const data = await resp.json();

                if (!resp.ok) throw new Error(data.error);

                document.getElementById('count-unreviewed').textContent = data.total;

                // Get glossary version
                const vresp = await fetch('/prionvault/api/glossary/version');
                const vdata = await vresp.json();
                document.getElementById('glossary-version').textContent = vdata.version;

                if (data.articles.length === 0) {
                    document.getElementById('article-list').innerHTML = '<div style="padding: 20px; text-align: center; color: #6b7280;">✓ No hay artículos sin procesar</div>';
                    return;
                }

                const html = data.articles.map(a => `
                    <div class="article-item" id="article-${a.id}">
                        <div class="article-info">
                            <div class="article-title">${escapeHtml(a.title)}</div>
                            <div class="article-meta">${a.summary_length} chars • ${a.authors || '—'}</div>
                        </div>
                        <span class="version-badge version-null" id="badge-${a.id}">NULL</span>
                        <div class="action-buttons">
                            <button class="btn-regenerate" onclick="regenerate('${a.id}')" id="btn-regen-${a.id}">🚀 Regenerar</button>
                            <button class="btn-check" onclick="checkStatus('${a.id}')" id="btn-check-${a.id}">✓ Verificar</button>
                        </div>
                    </div>
                `).join('');

                document.getElementById('article-list').innerHTML = html;
            } catch (e) {
                document.getElementById('alerts').innerHTML = `<div class="alert alert-error">❌ Error: ${e.message}</div>`;
            }
        }

        async function regenerate(articleId) {
            const btn = document.getElementById(`btn-regen-${articleId}`);
            const badge = document.getElementById(`badge-${articleId}`);

            btn.disabled = true;
            btn.innerHTML = '<span class="loading"></span> Regenerando...';
            badge.textContent = '⏳';

            try {
                const resp = await fetch(`/prionvault/api/glossary/regenerate-summary/${articleId}`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' }
                });

                const data = await resp.json();

                if (!resp.ok) throw new Error(data.error);

                document.getElementById('alerts').innerHTML = `<div class="alert alert-success">✅ Artículo regenerado. Verificando estado...</div>`;

                // Verificar automáticamente después de 2 segundos
                setTimeout(() => checkStatus(articleId), 2000);
            } catch (e) {
                badge.textContent = '❌';
                btn.disabled = false;
                btn.innerHTML = '🚀 Regenerar';
                document.getElementById('alerts').innerHTML = `<div class="alert alert-error">❌ Error en regeneración: ${e.message}</div>`;
            }
        }

        async function checkStatus(articleId) {
            const badge = document.getElementById(`badge-${articleId}`);
            const btn = document.getElementById(`btn-check-${articleId}`);

            btn.disabled = true;
            btn.innerHTML = '<span class="loading"></span>';

            try {
                const resp = await fetch(`/prionvault/api/glossary/diagnose-article/${articleId}`);
                const data = await resp.json();

                if (!resp.ok) throw new Error(data.error);

                const art = data.article;
                const version = art.ai_summary_glossary_version;

                if (version === null) {
                    badge.className = 'version-badge version-null';
                    badge.textContent = 'NULL ❌';
                    document.getElementById('alerts').innerHTML = `<div class="alert alert-error">❌ ai_summary_glossary_version sigue siendo NULL - El UPDATE no funcionó</div>`;
                } else {
                    badge.className = 'version-badge version-set';
                    badge.textContent = `v${version} ✓`;
                    document.getElementById('alerts').innerHTML = `<div class="alert alert-success">✅ ai_summary_glossary_version = ${version} - ¡Actualizado correctamente!</div>`;

                    // Recargar lista en 3 segundos
                    setTimeout(() => {
                        document.getElementById(`article-${articleId}`).style.opacity = '0.5';
                        document.getElementById(`article-${articleId}`).style.textDecoration = 'line-through';
                    }, 500);
                }

                btn.disabled = false;
                btn.innerHTML = '✓ Verificar';
            } catch (e) {
                btn.disabled = false;
                btn.innerHTML = '✓ Verificar';
                document.getElementById('alerts').innerHTML = `<div class="alert alert-error">❌ Error en verificación: ${e.message}</div>`;
            }
        }

        function escapeHtml(text) {
            const div = document.createElement('div');
            div.textContent = text;
            return div.innerHTML;
        }

        // Cargar al iniciar
        loadArticles();

        // Recargar cada 30 segundos para ver cambios
        setInterval(loadArticles, 30000);
        </script>
    </body>
    </html>
    """
    return Response(html, mimetype='text/html')


@prionvault_bp.route("/glossary/test-single", methods=["GET"])
@admin_required
def glossary_test_single_page():
    """Page to test single article improvement and regeneration."""
    html = """
    <h1>🧪 Prueba de Regeneración de Resumen</h1>
    <style>
    body { font-family: sans-serif; margin: 20px; }
    .section { margin: 20px 0; padding: 15px; border: 1px solid #ddd; border-radius: 4px; }
    .section h2 { margin-top: 0; }
    button { padding: 10px 20px; font-size: 14px; background: #0066cc; color: white; border: none; cursor: pointer; border-radius: 4px; margin: 5px; }
    button:hover { background: #0052a3; }
    button:disabled { background: #ccc; cursor: not-allowed; }
    input { padding: 8px; font-size: 14px; border: 1px solid #ddd; border-radius: 4px; width: 300px; }
    #result { margin-top: 15px; padding: 15px; background: #f5f5f5; border-radius: 4px; white-space: pre-wrap; font-family: monospace; font-size: 12px; max-height: 400px; overflow-y: auto; }
    #loading { display: none; color: #666; margin-top: 10px; }
    .ok { color: green; font-weight: bold; }
    .error { color: red; font-weight: bold; }
    .warning { color: orange; font-weight: bold; }
    table { border-collapse: collapse; width: 100%; margin: 10px 0; }
    table td, table th { border: 1px solid #ddd; padding: 8px; text-align: left; }
    table th { background: #f0f0f0; font-weight: bold; }
    </style>

    <div class="section">
      <h2>📋 Diagnóstico de Artículo</h2>
      <p>Ingresa el ID de un artículo para ver su estado:</p>
      <input type="text" id="articleId" placeholder="UUID del artículo (ej: 123e4567-e89b-12d3-a456-426614174000)" />
      <button onclick="diagnoseArticle()">🔍 Diagnosticar</button>
      <div id="diagResult"></div>
    </div>

    <div class="section">
      <h2>🔄 Regenerar Resumen</h2>
      <p>Regenera el resumen de un artículo específico:</p>
      <input type="text" id="regenerateId" placeholder="UUID del artículo" value="" />
      <button onclick="testRegenerate()">🚀 Regenerar Resumen</button>
      <div id="loading" style="display:none">⏳ Procesando... esto puede tomar 30-60 segundos</div>
      <div id="result"></div>
    </div>

    <script>
    async function diagnoseArticle() {
      const articleId = document.getElementById('articleId').value.trim();
      const result = document.getElementById('diagResult');

      if (!articleId) {
        result.innerHTML = '<span class="error">❌ Por favor ingresa un article ID</span>';
        return;
      }

      result.innerHTML = '<span class="warning">⏳ Diagnosticando...</span>';

      try {
        const res = await fetch(`/prionvault/api/glossary/diagnose-article/${articleId}`);
        const data = await res.json();

        if (!res.ok) {
          result.innerHTML = `<span class="error">❌ Error: ${data.error}</span>`;
          return;
        }

        const art = data.article;
        const html = `
<span class="ok">✓ Artículo encontrado</span>
<table>
  <tr><th>Campo</th><th>Valor</th></tr>
  <tr><td>ID</td><td>${art.id}</td></tr>
  <tr><td>Título</td><td>${art.title}</td></tr>
  <tr><td>Tiene Resumen</td><td>${art.has_summary ? '✓ Sí' : '✗ No'}</td></tr>
  <tr><td>Longitud Resumen</td><td>${art.summary_length} caracteres</td></tr>
  <tr><td>ai_summary_glossary_version</td><td><strong>${art.ai_summary_glossary_version || 'NULL (no procesado)'}</strong></td></tr>
  <tr><td>Versión Actual Glosario</td><td>${data.current_glossary_version}</td></tr>
  <tr><td>¿Necesita Procesamiento?</td><td>${art.has_summary && data.needs_processing ? '<span class="error">✓ SÍ</span>' : '✗ No'}</td></tr>
  <tr><td>¿Está Desactualizado?</td><td>${data.is_outdated ? '<span class="warning">✓ SÍ</span>' : '✗ No'}</td></tr>
  <tr><td>Actualizado</td><td>${new Date(art.updated_at).toLocaleString()}</td></tr>
</table>
        `;
        result.innerHTML = html;
      } catch (e) {
        result.innerHTML = `<span class="error">❌ Error de conexión: ${e.message}</span>`;
      }
    }

    async function testRegenerate() {
      const articleId = document.getElementById('regenerateId').value.trim();
      const loading = document.getElementById('loading');
      const result = document.getElementById('result');

      if (!articleId) {
        result.innerHTML = '<span class="error">❌ Por favor ingresa un article ID</span>';
        return;
      }

      loading.style.display = 'block';
      result.innerHTML = '';

      try {
        console.log('🚀 Regenerating article:', articleId);
        const res = await fetch(`/prionvault/api/glossary/regenerate-summary/${articleId}`, {
          method: 'POST',
          credentials: 'same-origin',
          headers: { 'Content-Type': 'application/json' }
        });

        const data = await res.json();
        console.log('📦 Response:', data);

        if (res.ok && data.ok) {
          result.innerHTML = `<span class="ok">✅ Regeneración Exitosa</span>
Artículo: ${data.article_id}
Versión Glosario: ${data.glossary_version}
Longitud Nuevo Resumen: ${data.new_summary_length} chars
Modelo: ${data.model_used}
Tokens: ${data.tokens_used}

Ahora diagnostica para verificar que ai_summary_glossary_version se actualizó:`;
          document.getElementById('articleId').value = articleId;
          setTimeout(() => diagnoseArticle(), 1500);
        } else {
          result.innerHTML = `<span class="error">❌ Error</span>
${JSON.stringify(data, null, 2)}`;
        }
      } catch (e) {
        console.error('Error:', e);
        result.innerHTML = `<span class="error">❌ Error de conexión</span>
${e.message}`;
      } finally {
        loading.style.display = 'none';
      }
    }

    // Auto-populate regenerateId from articleId when user types
    document.getElementById('articleId').addEventListener('change', () => {
      document.getElementById('regenerateId').value = document.getElementById('articleId').value;
    });
    </script>
    """
    return Response(html, mimetype='text/html')


@prionvault_bp.route("/glossary/diagnose", methods=["GET"])
@admin_required
def glossary_diagnose():
    """Diagnostic page for glossary processing."""
    import os
    from anthropic import Anthropic

    html = "<h1>Diagnóstico del Glosario</h1>"
    html += "<style>body{font-family:sans-serif;margin:20px}table{border-collapse:collapse;width:100%}td{border:1px solid #ccc;padding:10px}tr:nth-child(odd){background:#f9f9f9}.ok{color:green}.error{color:red}</style>"

    # Check API key
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    status = "✓" if api_key else "✗"
    html += f"<p><span class=\"{'ok' if api_key else 'error'}\">{status}</span> API Key configurada: {bool(api_key)}</p>"

    # Test Claude
    try:
        client = Anthropic(api_key=api_key) if api_key else None
        if client:
            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=50,
                messages=[{"role": "user", "content": "Say OK"}],
            )
            html += f"<p><span class=\"ok\">✓</span> Claude API funciona: {response.content[0].text if response.content else 'sin respuesta'}</p>"
        else:
            html += f"<p><span class=\"error\">✗</span> No se puede probar Claude sin API key</p>"
    except Exception as e:
        html += f"<p><span class=\"error\">✗</span> Error al conectar Claude: {str(e)}</p>"

    # Check database
    try:
        with db.engine.connect() as conn:
            count = conn.execute(sql_text("SELECT COUNT(*) FROM articles WHERE summary_ai IS NOT NULL")).scalar()
            html += f"<p><span class=\"ok\">✓</span> Base de datos OK: {count} artículos</p>"
    except Exception as e:
        html += f"<p><span class=\"error\">✗</span> Error de BD: {str(e)}</p>"

    # Batch status
    html += "<h2>Estado del Batch</h2>"
    html += f"<table>"
    html += f"<tr><td>Estado</td><td>{_batch_state.get('status', 'N/A')}</td></tr>"
    html += f"<tr><td>Procesados</td><td>{_batch_state.get('processed', 0)}</td></tr>"
    html += f"<tr><td>En cola</td><td>{_batch_state.get('queued', 0)}</td></tr>"
    if _batch_state.get('error'):
        html += f"<tr><td class=\"error\">Error</td><td class=\"error\">{_batch_state.get('error')}</td></tr>"
    html += "</table>"

    return Response(html, mimetype='text/html')


@prionvault_bp.route("/api/glossary/diagnose-article/<article_id>", methods=["GET"])
@admin_required
def api_glossary_diagnose_article(article_id):
    """Diagnose the state of a specific article."""
    try:
        with db.engine.connect() as conn:
            row = conn.execute(sql_text("""
                SELECT
                    id::text,
                    title,
                    summary_ai IS NOT NULL as has_summary,
                    COALESCE(char_length(summary_ai), 0) as summary_length,
                    ai_summary_glossary_version,
                    updated_at,
                    created_at
                FROM articles
                WHERE id = CAST(:aid AS UUID)
            """), {"aid": str(article_id)}).mappings().first()

        if not row:
            return jsonify({"error": "Article not found"}), 404

        row_dict = dict(row)
        current_version = None
        from .services import glossary_manager
        try:
            current_version = glossary_manager.get_current_glossary_version()
        except:
            pass

        return jsonify({
            "article": row_dict,
            "current_glossary_version": current_version,
            "needs_processing": row_dict['ai_summary_glossary_version'] is None if row_dict['has_summary'] else False,
            "is_outdated": row_dict['ai_summary_glossary_version'] is not None and current_version and row_dict['ai_summary_glossary_version'] < current_version if row_dict['ai_summary_glossary_version'] else None,
        })
    except Exception as e:
        logger.exception(f"Failed to diagnose article {article_id}")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/improve-batch", methods=["POST"])
@admin_required
def api_glossary_improve_batch():
    """Start a batch improvement run with glossary."""
    from .services import summary_improver, glossary_manager
    import uuid

    data = request.get_json(force=True, silent=True) or {}
    article_ids = data.get("article_ids", [])
    dry_run = data.get("dry_run", False)

    if not isinstance(article_ids, list) or not article_ids:
        return jsonify({"error": "article_ids must be a non-empty list"}), 400

    # Fetch current glossary
    try:
        glossary_context = glossary_manager.get_glossary_context()
        glossary_version = glossary_manager.get_current_glossary_version()
        if not glossary_context:
            return jsonify({"error": "No glossary terms available"}), 400
    except Exception as e:
        logger.exception("Failed to load glossary")
        return jsonify({"error": f"Glossary load failed: {str(e)[:200]}"}), 500

    # Generate batch ID
    batch_id = str(uuid.uuid4())

    # Reset batch state
    _batch_state["status"] = "processing"
    _batch_state["error"] = None
    _batch_state["queued"] = len(article_ids)
    _batch_state["processed"] = 0
    _batch_state["batch_id"] = batch_id
    _batch_state["current_article"] = ""
    _batch_state["current_status"] = "starting"

    # Run batch in background
    def _run():
        try:
            logger.info(f"🚀 Starting batch {batch_id} with {len(article_ids)} articles")
            _batch_state["current_status"] = "processing"

            def progress_callback(count, article_title="", status="processing"):
                _batch_state.update({
                    "processed": count,
                    "current_article": article_title,
                    "current_status": status,
                })
                logger.info(f"[PROGRESS] Batch {batch_id}: {count}/{len(article_ids)} - {article_title}")

            result = summary_improver.batch_improve_summaries(
                article_ids=article_ids,
                glossary_context=glossary_context,
                glossary_version=glossary_version,
                dry_run=dry_run,
                progress_callback=progress_callback,
            )

            _batch_state["status"] = "completed"
            logger.info(f"✅ Batch {batch_id} completed: {result}")
        except Exception as exc:
            logger.exception(f"❌ Batch {batch_id} failed: {exc}")
            _batch_state["status"] = "error"
            _batch_state["error"] = str(exc)[:200]

    threading.Thread(target=_run, name="pv-glossary-batch", daemon=True).start()

    return jsonify({
        "ok": True,
        "batch_id": batch_id,
        "queued": len(article_ids),
        "dry_run": dry_run,
        "glossary_version": glossary_version,
        "message": f"Queued {len(article_ids)} articles for improvement"
    })


# ── Glossary term operations ───────────────────────────────────────────────
@prionvault_bp.route("/api/glossary/terms", methods=["GET"])
@login_required
def api_glossary_terms():
    """Get all glossary terms, optionally filtered by category."""
    from .services import glossary_manager

    category = request.args.get("category", "")

    try:
        terms = glossary_manager.get_all_terms(category=category if category else None)
        return jsonify({
            "terms": terms,
            "count": len(terms),
        })
    except Exception as e:
        logger.exception("Failed to fetch glossary terms")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/term", methods=["PUT"])
@admin_required
def api_glossary_update_term():
    """Update a glossary term in-place."""
    from .services import glossary_manager

    data = request.get_json(force=True, silent=True) or {}
    term_en = (data.get("term_en") or "").strip().lower()
    term_es = (data.get("term_es_recommended") or "").strip()
    avoid = (data.get("term_es_avoid") or "").strip() or None
    notes = (data.get("notes") or "").strip() or None
    category = (data.get("category") or "").strip() or None
    version = data.get("version", 1)

    if not term_en or not term_es:
        return jsonify({"error": "term_en and term_es_recommended are required"}), 400

    try:
        result = glossary_manager.update_term(
            term_en=term_en,
            term_es_recommended=term_es,
            term_es_avoid=avoid,
            notes=notes,
            category=category,
            version=version
        )
        return jsonify({"ok": True, "updated": result})
    except Exception as e:
        logger.exception("Failed to update glossary term")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/categories", methods=["GET"])
@login_required
def api_glossary_categories():
    """Get all glossary categories."""
    from .services import glossary_manager

    try:
        categories = glossary_manager.get_categories()
        return jsonify({
            "categories": categories,
            "count": len(categories),
        })
    except Exception as e:
        logger.exception("Failed to fetch categories")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/version", methods=["GET"])
@login_required
def api_glossary_version():
    """Get current glossary version."""
    from .services import glossary_manager

    try:
        version = glossary_manager.get_current_glossary_version()
        return jsonify({"version": version})
    except Exception as e:
        logger.exception("Failed to fetch glossary version")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/import", methods=["POST"])
@admin_required
def api_glossary_import():
    """Import glossary terms from JSON or TSV.

    Accepts either:
    - JSON: {"terms": [{...}, ...]}
    - TSV: {"tsv_content": "English\\tCastellano...\\n..."}
    """
    from .services import glossary_manager

    data = request.get_json(force=True, silent=True) or {}

    # Try JSON format first
    terms = data.get("terms", [])
    if terms and isinstance(terms, list):
        try:
            result = glossary_manager.import_glossary(terms)
            return jsonify(result.__dict__ if hasattr(result, '__dict__') else result)
        except Exception as e:
            logger.exception("Glossary import failed")
            return jsonify({"error": str(e)[:300]}), 500

    # Try TSV format
    tsv_content = data.get("tsv_content", "")
    if tsv_content:
        try:
            # Validate first
            is_valid, errors, preview_rows = glossary_manager.validate_tsv_format(tsv_content)
            if not is_valid:
                return jsonify({"error": "TSV validation failed", "details": errors}), 400

            # Parse and import
            terms = glossary_manager.parse_tsv_to_terms(tsv_content)
            result = glossary_manager.import_glossary(terms)
            return jsonify(result.__dict__ if hasattr(result, '__dict__') else result)
        except Exception as e:
            logger.exception("TSV import failed")
            return jsonify({"error": str(e)[:300]}), 500

    return jsonify({"error": "Either 'terms' (JSON) or 'tsv_content' (TSV) is required"}), 400


# ── Excel export ───────────────────────────────────────────────────────────
@prionvault_bp.route("/api/glossary/export", methods=["GET"])
@admin_required
def api_glossary_export():
    """Export glossary improvement statistics as Excel file."""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from .services import summary_improver

        stats = summary_improver.get_improvement_stats()

        wb = Workbook()
        ws = wb.active
        ws.title = "Estadísticas"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Title
        ws['A1'] = "Estadísticas de mejora de resúmenes con glosario"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Summary metrics
        row = 3
        ws[f'A{row}'] = "Métrica"
        ws[f'B{row}'] = "Valor"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].fill = header_fill
        ws[f'B{row}'].font = header_font

        row += 1
        metrics = [
            ("Artículos mejorados", stats.get("total_articles_improved", 0)),
            ("Total de cambios", stats.get("total_changes", 0)),
            ("Cambios promedio/artículo", f"{stats.get('avg_changes_per_article', 0.0):.2f}"),
            ("Lotes procesados", stats.get("total_batches", 0)),
            ("Versión actual del glosario", stats.get("current_glossary_version", 0)),
            ("Última mejora", stats.get("last_improvement_at", "N/A")),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # By-version breakdown
        row += 2
        ws[f'A{row}'] = "Versión del glosario"
        ws[f'B{row}'] = "Artículos"
        ws[f'C{row}'] = "Cambios totales"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font

        row += 1
        for v in stats.get("by_version", []):
            ws[f'A{row}'] = v["glossary_version"]
            ws[f'B{row}'] = v["articles_improved"]
            ws[f'C{row}'] = v["total_changes"]
            row += 1

        # Most common corrections
        row += 2
        ws[f'A{row}'] = "Término original"
        ws[f'B{row}'] = "Término corregido"
        ws[f'C{row}'] = "Frecuencia"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font

        row += 1
        for corr in stats.get("most_common_corrections", []):
            ws[f'A{row}'] = corr["original"]
            ws[f'B{row}'] = corr["corrected"]
            ws[f'C{row}'] = corr["frequency"]
            row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=estadisticas_glosario.xlsx"}
        )

    except Exception as e:
        logger.exception(f"Failed to export statistics: {e}")
        return jsonify({"error": str(e)[:300]}), 500


# ── Cost tracking ──────────────────────────────────────────────────────────
@prionvault_bp.route("/api/glossary/batch/cost/<batch_id>", methods=["GET"])
@admin_required
def api_batch_cost(batch_id):
    """Get estimated cost for a specific batch improvement.

    Returns:
      {
        "batch_id": "uuid",
        "articles_processed": 100,
        "estimated_cost_eur": 0.05,
        "estimated_cost_usd": 0.055,
        "note": "Estimation based on ~€0.0005 per article",
        "timestamp": "2026-07-19T12:34:56"
      }
    """
    try:
        with db.engine.connect() as conn:
            result = conn.execute(sql_text("""
                SELECT
                  COUNT(DISTINCT article_id) as articles,
                  MAX(improved_at) as timestamp
                FROM summary_improvement_log
                WHERE batch_id = :batch_id AND dry_run = FALSE
            """), {"batch_id": batch_id}).first()

            if not result or result[0] == 0:
                return jsonify({"error": "Batch not found"}), 404

            articles, timestamp = result
            # Claude Haiku cost: ~€0.00015-0.0002 per article for terminology replacement
            est_cost_eur = articles * 0.00015
            est_cost_usd = est_cost_eur * 1.10

            return jsonify({
                "batch_id": batch_id,
                "articles_processed": int(articles),
                "estimated_cost_eur": round(est_cost_eur, 4),
                "estimated_cost_usd": round(est_cost_usd, 4),
                "cost_summary": f"~€{round(est_cost_eur, 4)} (Claude Haiku - terminology replacement only)",
                "note": "Cost reflects Claude Haiku API calls for exact term replacement (no regeneration)",
                "timestamp": str(timestamp) if timestamp else None,
            })

    except Exception as e:
        logger.exception(f"Failed to fetch batch cost for {batch_id}")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/costs/summary", methods=["GET"])
@admin_required
def api_costs_summary():
    """Get estimated cost summary for glossary improvements.

    Query params:
      - days: Number of days to look back (default: 30)
      - limit: Max batches to return (default: 10)

    Returns estimated costs based on articles processed.
    """
    days = request.args.get("days", 30, type=int)
    limit = request.args.get("limit", 10, type=int)

    try:
        with db.engine.connect() as conn:
            # Get summary stats
            summary = conn.execute(sql_text("""
                SELECT
                  COUNT(DISTINCT batch_id) as batch_count,
                  COUNT(DISTINCT article_id) as article_count
                FROM summary_improvement_log
                WHERE dry_run = FALSE
                  AND improved_at >= NOW() - INTERVAL '1 day' * :days
            """), {"days": days}).first()

            batch_count, article_count = summary

            # Claude Haiku cost: ~€0.00015 per article
            est_total_eur = article_count * 0.00015
            est_total_usd = est_total_eur * 1.10
            avg_cost_per_article = 0.00015

            # Get recent batches
            batches = conn.execute(sql_text("""
                SELECT
                  batch_id,
                  COUNT(DISTINCT article_id) as articles,
                  MAX(improved_at) as timestamp
                FROM summary_improvement_log
                WHERE dry_run = FALSE
                  AND improved_at >= NOW() - INTERVAL '1 day' * :days
                GROUP BY batch_id
                ORDER BY timestamp DESC
                LIMIT :lim
            """), {"days": days, "lim": limit}).fetchall()

            recent_batches = []
            for batch_id, articles, ts in batches:
                batch_est_eur = articles * 0.00015
                batch_est_usd = batch_est_eur * 1.10
                recent_batches.append({
                    "batch_id": batch_id,
                    "articles": int(articles),
                    "estimated_cost_eur": round(batch_est_eur, 4),
                    "estimated_cost_usd": round(batch_est_usd, 4),
                    "timestamp": str(ts) if ts else None,
                })

            return jsonify({
                "period_days": days,
                "total_batches": int(batch_count),
                "total_articles": int(article_count),
                "estimated_total_eur": round(est_total_eur, 4),
                "estimated_total_usd": round(est_total_usd, 4),
                "avg_estimated_cost_per_article": round(avg_cost_per_article, 6),
                "note": "Cost reflects Claude Haiku API for exact terminology replacement (structure preserved)",
                "recent_batches": recent_batches,
            })

    except Exception as e:
        logger.exception("Failed to fetch cost summary")
        return jsonify({"error": str(e)[:300]}), 500


# ── Public glossary API (for other modules: PrionRead, PrionPacks, etc.) ─────
@prionvault_bp.route("/api/glossary/public/version", methods=["GET"])
@login_required
def api_glossary_public_version():
    """Get current glossary version (public API for other modules)."""
    from .services import glossary_manager
    try:
        version = glossary_manager.get_current_glossary_version()
        return jsonify({"version": version})
    except Exception as e:
        logger.exception("Failed to fetch glossary version")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/public/terms", methods=["GET"])
@login_required
def api_glossary_public_terms():
    """Get all glossary terms (public API for other modules).

    Optional query params:
    - category: Filter by category
    - limit: Max results (default 1000)
    - offset: Pagination offset (default 0)
    """
    from .services import glossary_manager

    try:
        category = request.args.get("category", "").strip() or None
        limit = min(int(request.args.get("limit", 1000)), 5000)
        offset = int(request.args.get("offset", 0))

        terms = glossary_manager.get_all_terms(category=category)
        total = len(terms)
        paginated = terms[offset:offset + limit]

        return jsonify({
            "version": glossary_manager.get_current_glossary_version(),
            "total": total,
            "returned": len(paginated),
            "offset": offset,
            "limit": limit,
            "terms": [
                {
                    "term_en": t.get("term_en", "").lower(),
                    "term_es_recommended": t.get("term_es_recommended", ""),
                    "term_es_avoid": t.get("term_es_avoid"),
                    "category": t.get("category"),
                    "notes": t.get("notes"),
                }
                for t in paginated
            ]
        })
    except Exception as e:
        logger.exception("Failed to fetch glossary terms")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/recover-historical", methods=["POST"])
@admin_required
def api_glossary_recover_historical():
    """Recover improvement logs for articles improved before summary_improvement_log existed.

    Articles with ai_summary_glossary_version set but no entry in summary_improvement_log
    will be given a 'recovered' entry to maintain accurate tracking.
    """
    try:
        with db.engine.begin() as conn:
            # Find orphaned articles
            orphaned = conn.execute(sql_text("""
                SELECT COUNT(*) FROM articles a
                WHERE a.ai_summary_glossary_version IS NOT NULL
                  AND a.id NOT IN (
                    SELECT DISTINCT article_id FROM summary_improvement_log
                  )
            """)).scalar() or 0

            if orphaned == 0:
                return jsonify({
                    "ok": True,
                    "recovered": 0,
                    "message": "No historical articles to recover"
                })

            # Create recovery entries
            conn.execute(sql_text("""
                INSERT INTO summary_improvement_log (
                    article_id,
                    glossary_version_used,
                    original_summary,
                    improved_summary,
                    changes_count,
                    batch_id,
                    improved_at,
                    dry_run,
                    input_tokens,
                    output_tokens,
                    model_used
                )
                SELECT
                    a.id,
                    a.ai_summary_glossary_version,
                    a.summary_ai,
                    a.summary_ai,
                    0,
                    'recovered-' || gen_random_uuid()::text,
                    a.updated_at,
                    FALSE,
                    0,
                    0,
                    'recovered-from-field'
                FROM articles a
                WHERE a.ai_summary_glossary_version IS NOT NULL
                  AND a.id NOT IN (
                    SELECT DISTINCT article_id FROM summary_improvement_log
                  )
                ON CONFLICT DO NOTHING
            """))

        logger.info(f"✅ Recovered {orphaned} historical glossary improvements")
        return jsonify({
            "ok": True,
            "recovered": orphaned,
            "message": f"Recovered {orphaned} articles from historical improvements"
        })

    except Exception as e:
        logger.exception("Failed to recover historical improvements")
        return jsonify({"error": str(e)[:300]}), 500


@prionvault_bp.route("/api/glossary/public/search", methods=["GET"])
@login_required
def api_glossary_public_search():
    """Search glossary terms by English or Spanish (public API).

    Query params:
    - q: Search query (required)
    - limit: Max results (default 50)
    """
    from .services import glossary_manager

    try:
        query = request.args.get("q", "").strip().lower()
        if not query:
            return jsonify({"error": "Missing 'q' parameter"}), 400

        limit = min(int(request.args.get("limit", 50)), 500)

        all_terms = glossary_manager.get_all_terms()
        matches = []

        for term in all_terms:
            term_en = (term.get("term_en") or "").lower()
            term_es = (term.get("term_es_recommended") or "").lower()
            avoid = (term.get("term_es_avoid") or "").lower()

            # Simple substring match (can be improved with fuzzy matching)
            if query in term_en or query in term_es or (avoid and query in avoid):
                matches.append({
                    "term_en": term.get("term_en", ""),
                    "term_es_recommended": term.get("term_es_recommended", ""),
                    "term_es_avoid": term.get("term_es_avoid"),
                    "category": term.get("category"),
                    "notes": term.get("notes"),
                })

            if len(matches) >= limit:
                break

        return jsonify({
            "query": query,
            "found": len(matches),
            "limited_to": limit if len(matches) >= limit else None,
            "results": matches[:limit]
        })
    except Exception as e:
        logger.exception("Failed to search glossary")
        return jsonify({"error": str(e)[:300]}), 500


# ── Regenerate complete summary with glossary ──────────────────────────────
@prionvault_bp.route("/api/glossary/regenerate-summary/<article_id>", methods=["POST"])
@admin_required
def api_glossary_regenerate_summary(article_id):
    """Regenerate a complete AI summary for an article with glossary applied.

    This generates a brand new summary from scratch (using title, authors, year, etc.)
    rather than just improving the existing one. The new summary automatically
    includes glossary terminology.

    After successful regeneration, marks the article as processed by current
    glossary version so it won't appear in "Choose articles to improve" list.
    """
    from .services import ai_summary, glossary_manager

    try:
        logger.info(f"📖 Starting regenerate for article {article_id}")
        glossary_version = glossary_manager.get_current_glossary_version()
        logger.info(f"🔖 Using glossary version {glossary_version}")

        # Fetch article metadata
        with db.engine.connect() as conn:
            row = conn.execute(sql_text(
                """SELECT id, title, authors, year, journal, doi, pubmed_id, extracted_text
                   FROM articles WHERE id = CAST(:aid AS UUID)"""
            ), {"aid": str(article_id)}).first()

        if not row:
            logger.error(f"❌ Article not found: {article_id}")
            return jsonify({"error": "Article not found"}), 404

        article_id_val, title, authors, year, journal, doi, pubmed_id, extracted_text = row
        logger.info(f"✓ Found article: {title[:50]}")

        # Generate new summary with glossary (automatically applied in system prompt)
        logger.info(f"🤖 Calling generate_summary for article {article_id}")
        result = ai_summary.generate_summary(
            title=title,
            authors=authors,
            year=year,
            journal=journal,
            doi=doi,
            pubmed_id=pubmed_id,
            extracted_text=extracted_text,
        )
        logger.info(f"✓ Generated summary: {len(result.text)} chars")

        # Save regenerated summary to database
        logger.info(f"💾 Updating database with new summary and glossary_version={glossary_version}")
        with db.engine.begin() as conn:
            update_result = conn.execute(sql_text(
                """UPDATE articles
                   SET summary_ai = :summary,
                       ai_summary_glossary_version = :ver,
                       updated_at = NOW()
                   WHERE id = CAST(:aid AS UUID)"""
            ), {
                "summary": result.text,
                "ver": glossary_version,
                "aid": str(article_id),
            })
            logger.info(f"✓ Updated {update_result.rowcount} rows in articles table")

        # Verify the update
        with db.engine.connect() as conn:
            verify = conn.execute(sql_text(
                """SELECT ai_summary_glossary_version FROM articles WHERE id = CAST(:aid AS UUID)"""
            ), {"aid": str(article_id)}).scalar()
            logger.info(f"✓ Verified: ai_summary_glossary_version = {verify}")

        logger.info(f"✅ Successfully regenerated summary for article {article_id}")

        return jsonify({
            "ok": True,
            "article_id": str(article_id),
            "glossary_version": glossary_version,
            "new_summary_length": len(result.text),
            "model_used": result.model,
            "tokens_used": (result.tokens_in or 0) + (result.tokens_out or 0),
            "message": "Summary regenerated and glossary applied successfully"
        })

    except ai_summary.NotConfigured as e:
        logger.error(f"❌ AI provider not configured: {e}")
        return jsonify({"error": f"AI provider not available: {str(e)[:200]}"}), 503
    except Exception as e:
        logger.exception(f"❌ Failed to regenerate summary for {article_id}: {e}")
        return jsonify({"error": str(e)[:300]}), 500
